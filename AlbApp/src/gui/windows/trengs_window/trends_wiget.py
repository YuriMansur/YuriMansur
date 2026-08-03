import datetime as _dt                          # форматирование меток времени на графике
import numpy as np                              # кольцевые буферы и операции с массивами точек
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,   # базовые виджеты и компоновка
    QColorDialog, QFileDialog,                        # диалоги выбора цвета и сохранения файла
    QFrame, QDateEdit, QTimeEdit,                     # рамка навигации, поля ввода даты/времени
    QLabel, QSpinBox, QComboBox,
    QSlider, QWidgetAction, QCheckBox, QMenu,         # слайдер прозрачности, меню строки
    QSizePolicy,                                      # панель каналов — по содержимому
    QDialog, QDialogButtonBox,                        # диалог выбора каналов для выгрузки
)
from PyQt6.QtCore import Qt, QDateTime, QTimer, QSize  # флаги Qt, дата/время, таймер, размер значка
from PyQt6.QtGui import (
    QShortcut, QKeySequence, QAction, QColor, QIcon, QPalette,  # клавиши, меню, цвет, значок, палитра
)
from PyQt6.QtWidgets import QApplication      # палитра приложения → цвета панелей

from gui.icons import make_icon                 # значки панели: глаз, режимы
import pyqtgraph as pg                          # графическая библиотека (PlotWidget, InfiniteLine и др.)
pg.setConfigOptions(antialias=True, useOpenGL=True)
from event_bus import bus      # шина событий: получение точек от worker'а

# ── локальные модули ──────────────────────────────────────────────────────────
from ._archive_worker import (
    INFLUX_BUCKET,       # имя bucket в InfluxDB
    LIVE_RENDER_MS,      # интервал таймера скролла live-графика (мс)
    LIVE_WINDOW_SECS,    # глубина отображаемого live-окна (сек)
    LIVE_POINTS_START,   # стартовая ёмкость live-буфера (темп ещё не известен)
    LIVE_POINTS_LIMIT,   # потолок ёмкости live-буфера
    LIVE_POINTS_MARGIN,  # запас ёмкости сверх окна
    N_ARCHIVE_WORKERS,   # количество параллельных потоков загрузки архива
    _ArchiveWorker,      # QThread-воркер одного параллельного запроса к InfluxDB
)
from ._axis_item import _TimeAxisItem           # кастомная ось X с метками ЧЧ:ММ:СС
from ._pg_menu_utils import (
    _translate_pg_menus,                        # русификация контекстного меню pyqtgraph
    _install_x_time_format,                     # форматирование полей диапазона оси X как время
)

# цвета каналов по умолчанию
_CH_COLORS = ["#e67e22", "#3498db", "#2ecc71"]

# толщина линии: одна на все каналы, из панели не настраивается — выбор убран,
# менять можно только программно через _set_ch_width
_CH_WIDTH = 2

# имена каналов (индекс 0 = канал 1)
_CH_NAMES = [
    "Текущая уставка нагружения",
    "Датчик нагружения",
    "Датчик перемещения",
]

_ARROW_CACHE: dict[str, str] = {}


def _arrow_url(kind: str, color: str) -> str:
    """Путь к PNG-стрелке для QSpinBox.

    Стрелки счётчика stylesheet умеет задавать только картинкой: как только у
    ::up-button появляется свой фон, штатные Qt рисовать перестаёт, а нарисовать
    треугольник правилами CSS в Qt нечем. Поэтому берём тот же значок, что и
    везде в приложении, и один раз кладём его в кэш-каталог.
    """
    key = f"{kind}_{color.lstrip('#')}"
    path = _ARROW_CACHE.get(key)
    if path is None:
        import os, tempfile
        d = os.path.join(tempfile.gettempdir(), "albapp_icons")
        os.makedirs(d, exist_ok=True)
        path = os.path.join(d, key + ".png").replace("\\", "/")
        make_icon(kind, color, 16).save(path)
        _ARROW_CACHE[key] = path
    return path


def _theme(dark: bool) -> tuple[str, str, str, str, str]:
    """Цвета панелей трендов: (фон, текст, фон элементов, рамка, углубление).

    Берём из палитры приложения, а не своим набором: раньше панели были
    сине-серые (#2c3e50), и вкладка трендов выбивалась из остального окна.
    Если палитра ещё не выставлена (тесты, отдельный запуск виджета) —
    подставляем те же значения, что задаёт главное окно.
    """
    fallback = (("#282828", "#e0e0e0", "#2d2d2d", "#3c3c3c", "#161616") if dark
                else ("#e9e9e9", "#141414", "#e1e1e1", "#b4b4b4", "#ffffff"))
    app = QApplication.instance()
    pal = app.palette() if app is not None else None
    R = QPalette.ColorRole
    # палитра могла ещё не смениться на нужную тему — тогда берём запасную
    if pal is None or (pal.color(R.Window).lightness() < 128) != dark:
        return fallback
    return (pal.color(R.AlternateBase).name(),
            pal.color(R.WindowText).name(),
            pal.color(R.Button).name(),
            pal.color(R.Mid).name(),
            pal.color(R.Base).name())


def _panel_style(dark: bool) -> str:
    bg, text, ctrl_bg, border, _base = _theme(dark)
    return f"""
    QFrame#chPanel {{
        background-color: {bg};
    }}
    QWidget {{
        background-color: {bg};
    }}
    QLabel {{
        color: {text};
        background: transparent;
    }}
    QSpinBox {{
        background: {ctrl_bg};
        color: {text};
        border: 1px solid {border};
        border-radius: 3px;
        padding: 1px 4px;
        min-height: 20px;
    }}
    QSpinBox::up-button, QSpinBox::down-button {{
        background: {border};
        border: none;
        width: 14px;
    }}
    QSpinBox::up-button:hover, QSpinBox::down-button:hover {{
        background: #3498db;
    }}
    QSpinBox::up-arrow {{
        image: url({_arrow_url("arrow_up", text)});
        width: 8px; height: 8px;
    }}
    QSpinBox::down-arrow {{
        image: url({_arrow_url("arrow_down", text)});
        width: 8px; height: 8px;
    }}
    QCheckBox {{
        color: {text};
        background: transparent;
        spacing: 5px;
    }}
    QCheckBox::indicator {{
        width: 14px;
        height: 14px;
        border: 1px solid {border};
        border-radius: 2px;
        background: {ctrl_bg};
    }}
    QCheckBox::indicator:checked {{
        background: #3498db;
        border-color: #2980b9;
    }}
    QSlider::groove:horizontal {{
        background: {border};
        height: 4px;
        border-radius: 2px;
    }}
    QSlider::handle:horizontal {{
        background: #3498db;
        width: 12px;
        height: 12px;
        margin: -4px 0;
        border-radius: 6px;
        border: none;
    }}
    QSlider::handle:horizontal:hover {{
        background: #5dade2;
    }}
"""


# цвета режимов: live — «идёт сейчас», архив — «поднято из базы»
_MODE_COLORS = {"live": "#2ecc71", "archive": "#3498db"}


def _nav_style(dark: bool) -> str:
    """Панель режимов: подпись, сегментный переключатель, кнопки Live/Архив.

    Выбранный сегмент заливается цветом своего режима, невыбранный остаётся
    плоским — оператор видит текущий режим, не читая надписей.
    """
    bg, text, _ctrl, seg_brd, seg_bg = _theme(dark)
    return f"""
    QFrame {{ background-color: {bg}; }}
    QFrame#modeSeg {{
        background: {seg_bg}; border: 1px solid {seg_brd};
        border-bottom: 1px solid {seg_brd}; border-radius: 6px;
    }}
    QFrame#modeSeg QPushButton {{
        color: {text}; background: transparent; border: none;
        border-radius: 4px; padding: 4px 14px; font-size: 12px; min-width: 64px;
    }}
    QFrame#modeSeg QPushButton:hover:!checked {{
        background: {_rgba(text, 26)};
    }}
    QFrame#modeSeg QPushButton#mode_live:checked {{
        background: {_MODE_COLORS['live']}; color: #ffffff; font-weight: bold;
    }}
    QFrame#modeSeg QPushButton#mode_archive:checked {{
        background: {_MODE_COLORS['archive']}; color: #ffffff; font-weight: bold;
    }}
    """


class TrendsWiget(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        # {ch_id: dict} — 3 фиксированных канала; структура ch-словаря описана в _setup_ui
        self._channels: dict = {}

        self._mode                    = ""
        self._archive_lines: dict     = {}   # {ch_id: PlotDataItem}
        self._archive_parts: dict     = {}   # {ch_id: {worker_idx: (times, values)}}
        self._archive_workers: list   = []   # плоский список всех активных воркеров
        self._workers_done            = 0
        self._total_archive_workers   = 0
        self._arch_full:   tuple = None      # весь заказанный диапазон, сек Unix
        self._arch_window: tuple = None      # что реально загружено сейчас
        self._auto_scroll: bool  = True
        self._y_range:     tuple = (0.0, 1.0)
        self._ts_offset:   float = 0.0       # нормализация X для OpenGL (вычитается из timestamps)
        self._live_start_ts: float = 0.0     # момент включения live (левый край до заполнения окна)

        # Ёмкость кольцевых live-буферов. Стартовая — до того, как воркер сообщит
        # реальный шаг между точками; дальше пересчитывается под окно (см. _on_stream_rate).
        self._live_cap:       int  = LIVE_POINTS_START
        self._stream_step_ms: dict = {}      # имя потока → шаг между точками, мс
        bus.stream_rate.connect(self._on_stream_rate)

        # таймер плавного скролла X-оси
        self._render_timer = QTimer(self)
        self._render_timer.setTimerType(Qt.TimerType.PreciseTimer)
        self._render_timer.timeout.connect(self._render_frame)

        self._setup_ui()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(4, 4, 4, 4)
        root.setSpacing(4)

        # ── переключатель режимов ────────────────────────────────────────────
        self._nav_frame = QFrame()
        self._nav_frame.setStyleSheet(_nav_style(True))
        nav_frame = self._nav_frame
        row_mode = QHBoxLayout(nav_frame)
        row_mode.setContentsMargins(6, 4, 6, 4)
        row_mode.setSpacing(6)

        # Сегментный переключатель вместо выпадающего списка: режима всего два,
        # оба должны быть видны сразу — какой сейчас включён, читается без клика.
        self._mode_btns = {}
        seg = QFrame(); seg.setObjectName("modeSeg")
        seg_row = QHBoxLayout(seg)
        seg_row.setContentsMargins(2, 2, 2, 2); seg_row.setSpacing(2)
        for mode, text, kind in (("live", "Live", "record"), ("archive", "Архив", "doc")):
            btn = QPushButton(" " + text)
            btn.setObjectName(f"mode_{mode}")
            btn.setCheckable(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setIconSize(QSize(14, 14))
            btn.clicked.connect(lambda _c=False, m=mode: self._set_mode(m))
            seg_row.addWidget(btn)
            self._mode_btns[mode] = (btn, kind)
        row_mode.addWidget(seg)
        self._row_mode = row_mode      # сюда же встаёт панель архива (см. ниже)
        row_mode.addStretch()
        root.addWidget(nav_frame)

        # ── панель каналов ───────────────────────────────────────────────────
        self._ch_frame = QFrame()
        self._ch_frame.setObjectName("chPanel")
        self._ch_frame.setStyleSheet(_panel_style(True))
        ch_frame = self._ch_frame
        # строки каналов — в ряд: панель занимает одну строку по высоте и
        # оставляет графику больше места
        ch_outer = QHBoxLayout(ch_frame)
        ch_outer.setContentsMargins(4, 3, 4, 3)
        ch_outer.setSpacing(6)
        # панель по содержимому: расти вширь ей незачем, пустое место справа
        # только делало её похожей на недозаполненную таблицу
        ch_frame.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Maximum)
        # каналы и панель архива — одной строкой, архив справа от каналов
        self._controls_row = QHBoxLayout()
        self._controls_row.setContentsMargins(0, 0, 0, 0)
        self._controls_row.setSpacing(6)
        self._controls_row.addWidget(ch_frame, 0, Qt.AlignmentFlag.AlignTop)
        self._controls_row.addStretch()
        root.addLayout(self._controls_row)


        QShortcut(QKeySequence(Qt.Key.Key_Left),  self, self._pan_left)
        QShortcut(QKeySequence(Qt.Key.Key_Right), self, self._pan_right)

        # ── панель архива ────────────────────────────────────────────────────
        # Диапазон и кнопка загрузки в одну строку, справа от каналов: готовых
        # интервалов (5м/30м/1ч…) больше нет, границы задаются полями.
        self._arch_panel = QFrame()
        self._arch_panel.setObjectName("chPanel")
        self._arch_panel.setStyleSheet(_panel_style(True))
        self._arch_panel.setSizePolicy(QSizePolicy.Policy.Maximum,
                                       QSizePolicy.Policy.Maximum)
        arch_layout = QHBoxLayout(self._arch_panel)
        arch_layout.setContentsMargins(4, 3, 4, 3)
        arch_layout.setSpacing(3)

        _FLD_H = 22        # общая высота элементов строки

        now = QDateTime.currentDateTime()
        self.date_from = QDateEdit(now.addSecs(-300).date())
        self.date_from.setDisplayFormat("dd.MM.yyyy")
        self.date_from.setCalendarPopup(True)
        self.time_from = QTimeEdit(now.addSecs(-300).time())
        self.time_from.setDisplayFormat("HH:mm:ss")

        self.date_to = QDateEdit(now.date())
        self.date_to.setDisplayFormat("dd.MM.yyyy")
        self.date_to.setCalendarPopup(True)
        self.time_to = QTimeEdit(now.time())
        self.time_to.setDisplayFormat("HH:mm:ss")

        for field in (self.date_from, self.time_from, self.date_to, self.time_to):
            field.setFixedHeight(_FLD_H)

        self.btn_load = QPushButton("Загрузить")
        self.btn_load.setFixedHeight(_FLD_H)
        self.btn_load.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_load.clicked.connect(self._load_archive)

        arch_layout.addWidget(_cap("С"))
        arch_layout.addWidget(self.date_from)
        arch_layout.addWidget(self.time_from)
        arch_layout.addWidget(_cap("по"))
        arch_layout.addWidget(self.date_to)
        arch_layout.addWidget(self.time_to)
        arch_layout.addWidget(_sep())
        arch_layout.addWidget(self.btn_load)

        # Справа от каналов, в той же строке; распорка в конце — последняя.
        self._arch_panel.setVisible(False)
        self._controls_row.insertWidget(self._controls_row.count() - 1,
                                        self._arch_panel, 0,
                                        Qt.AlignmentFlag.AlignTop)

        # ── график ───────────────────────────────────────────────────────────
        self._time_axis  = _TimeAxisItem(orientation="bottom")
        self.plot_widget = pg.PlotWidget(
            axisItems={"bottom": self._time_axis})
        self.plot_widget.setBackground("#2c2c2c")
        self.plot_widget.showGrid(x=True, y=True, alpha=0.3)
        _auto_btn = self.plot_widget.getPlotItem().autoBtn
        _auto_btn.hide()
        _auto_btn.show = lambda: None  # запретить pyqtgraph показывать кнопку
        self._legend = self.plot_widget.addLegend()
        root.addWidget(self.plot_widget, 1)

        # перевод контекстного меню
        self._csv_export_action, self._xlsx_export_action, self._export_menu_action = \
            _translate_pg_menus(self.plot_widget)
        self._csv_export_action .triggered.connect(self._export_csv)
        self._xlsx_export_action.triggered.connect(self._export_excel)

        # поля ручного диапазона оси X → формат ЧЧ:ММ:СС
        _install_x_time_format(self.plot_widget)
        self._build_graph_context_menu()

        # ── перекрестие + метка значения ────────────────────────────────────
        dash = pg.mkPen(color="#888888", width=1, style=Qt.PenStyle.DashLine)
        self._vline = pg.InfiniteLine(angle=90, movable=False, pen=dash)
        self._hline = pg.InfiniteLine(angle=0,  movable=False, pen=dash)
        self._vline.setVisible(False)
        self._hline.setVisible(False)
        self.plot_widget.addItem(self._vline, ignoreBounds=True)
        self.plot_widget.addItem(self._hline, ignoreBounds=True)

        # ── маркер по клику ──────────────────────────────────────────────────
        self._click_marker = pg.ScatterPlotItem(
            size=10,
            pen=pg.mkPen("#cc0000", width=2),
            brush=pg.mkBrush(255, 80, 80, 200),
        )
        self._click_label = pg.TextItem(
            anchor=(0, 1), color="#aa0000",
            fill=pg.mkBrush(255, 230, 230, 220),
        )
        self._click_marker.setZValue(30)
        self._click_label.setZValue(30)
        self._click_marker.setVisible(False)
        self._click_label.setVisible(False)
        self.plot_widget.addItem(self._click_marker, ignoreBounds=True)
        self.plot_widget.addItem(self._click_label,  ignoreBounds=True)

        self.plot_widget.scene().sigMouseMoved.connect(self._on_mouse_moved)
        self.plot_widget.scene().sigMouseClicked.connect(self._on_mouse_clicked)
        self.plot_widget.getViewBox().sigRangeChangedManually.connect(self._on_manual_pan)

        # Архив перечитывается под видимое окно. Задержка — чтобы при
        # перетаскивании ушёл один запрос, а не по одному на каждый кадр.
        self._arch_timer = QTimer(self)
        self._arch_timer.setSingleShot(True)
        self._arch_timer.setInterval(300)
        self._arch_timer.timeout.connect(self._reload_visible_archive)
        self.plot_widget.getViewBox().sigXRangeChanged.connect(self._on_arch_x_changed)

        # Ctrl+колесо → масштаб оси X; обычное колесо → масштаб оси Y
        _vb = self.plot_widget.getViewBox()
        _vb_wheel = type(_vb).wheelEvent
        def _wheel(ev, axis=None):
            if ev.modifiers() & Qt.KeyboardModifier.ControlModifier:
                _vb_wheel(_vb, ev, axis=0)
            else:
                _vb_wheel(_vb, ev, axis=1)
        _vb.wheelEvent = _wheel

        # ── 3 фиксированных канала ────────────────────────────────────────────

        for ch_id, (color, name) in enumerate(zip(_CH_COLORS, _CH_NAMES), start=1):
            self._init_channel(ch_id, color, name)
            row = self._make_channel_row(ch_id)
            self._channels[ch_id]['row'] = row
            ch_outer.addWidget(row)

        # статическая привязка каналов к именам потоков (источник — bus.stream_points)
        # порядок combo: 0=не привязан, 1=Датчик нагружения, 2=Датчик перемещения, 3=Текущая уставка нагружения
        _static = [
            (1, 3, 'setpoint'),       # Текущая уставка нагружения
            (2, 1, 'tenza'),          # Датчик нагружения
            (3, 2, 'displacement'),   # Датчик перемещения
        ]
        for ch_id, _, stream in _static:
            self._channels[ch_id]['signal'] = bus.stream_points
            self._channels[ch_id]['stream'] = stream
        self._channels[1]['hold'] = True   # уставка тянется непрерывно до now
        self._channels[1]['curve'].setClipToView(False)  # hold-канал рисует шаг-функцию вручную

        self._set_mode("live")

    # ── инициализация каналов ─────────────────────────────────────────────────

    def _init_channel(self, ch_id: int, color: str, name: str = ""):
        """Создать запись канала и его кривую; строку UI не создаёт."""
        curve = self.plot_widget.plot(
            [], [],
            pen=pg.mkPen(color=QColor(color), width=_CH_WIDTH),
        )
        curve.setDownsampling(auto=True, method='peak')
        curve.setClipToView(True)   # для hold-каналов сбрасывается ниже в _setup_ui
        curve.setSkipFiniteCheck(True)  # данные всегда конечны — пропуск NaN/inf проверки

        self._channels[ch_id] = {
            'curve':        curve,
            'name':         name or f"канал {ch_id}",
            'buf_t':        np.empty(self._live_cap, dtype=np.float64),
            'buf_v':        np.empty(self._live_cap, dtype=np.float64),
            'write':        0,
            'full':         False,
            'color':        color,
            'width':        _CH_WIDTH,
            'alpha':        255,
            'points':       False,
            'visible':      True,   # отображается ли канал на графике
            'hold':         False,  # True → линия продлевается до now на каждом кадре
            'last_val':     None,   # последнее полученное значение (для hold-режима)
            'signal':       None,
            'stream':       None,   # имя потока в bus.stream_points (фильтр в слоте)
            'slot':         None,   # сохранённая ссылка на lambda-слот для отключения
            'row':          None,
        }

    def _make_channel_row(self, ch_id: int) -> QWidget:
        """Собрать горизонтальную строку управления каналом."""
        ch = self._channels[ch_id]
        row = QFrame()
        row.setObjectName("chRow")
        row.setFixedHeight(24)
        ch['row'] = row              # нужен уже сейчас: по нему красит _style_ch_row
        hl = QHBoxLayout(row)
        hl.setContentsMargins(6, 0, 6, 0)
        hl.setSpacing(6)

        # Отдельных кнопок «глаз» и «цвет» в строке нет — и то, и другое лежит в
        # меню по правой кнопке. Обычный щелчок ничего не делает намеренно:
        # случайно попасть по строке и потерять канал с графика не должно.
        # Что канал скрыт, видно по самой строке: она уходит в серое целиком.
        row.setToolTip("Правая кнопка — цвет линии и показ канала")
        row.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        row.customContextMenuRequested.connect(
            lambda pos, cid=ch_id: self._row_menu(cid, pos))

        # ── лейбл ────────────────────────────────────────────────────────────
        # имя цветом своей кривой: строку находишь по цвету, а не вчитываясь
        lbl = QLabel(ch['name'])
        lbl.setObjectName("chName")
        # ширина под самое длинное имя, а не «на глаз»: колонка ровная у всех строк
        lbl.setFixedWidth(max(lbl.fontMetrics().horizontalAdvance(n)
                              for n in _CH_NAMES) + 8)
        ch['name_lbl'] = lbl

        # ── прозрачность ─────────────────────────────────────────────────────
        alpha_slider = QSlider(Qt.Orientation.Horizontal)
        alpha_slider.setRange(0, 100)
        alpha_slider.setValue(100)
        # минимум небольшой: на узких экранах карточки должны ужаться, чтобы
        # панель архива справа осталась на той же строке
        alpha_slider.setMinimumWidth(45)
        alpha_slider.setToolTip("Прозрачность линии")
        alpha_slider.valueChanged.connect(
            lambda v, cid=ch_id: self._set_ch_alpha(cid, v))

        # ── точки ────────────────────────────────────────────────────────────
        # значком, а не галочкой: сама иконка показывает, что получится —
        # ломаная с маркерами или без них
        points_btn = QPushButton()
        points_btn.setObjectName("chPoints")
        points_btn.setCheckable(True)
        points_btn.setChecked(ch['points'])
        points_btn.setFixedSize(24, 20)
        points_btn.setIconSize(QSize(17, 17))
        points_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        points_btn.toggled.connect(
            lambda v, cid=ch_id: self._set_ch_points(cid, v))
        ch['points_btn'] = points_btn

        # Слабину забирает слайдер (stretch=1), распорки в конце нет — иначе
        # справа оставался пустой хвост.
        # Подписи «Прозрачность» нет: в ряд встают три карточки, и вместе с
        # панелью архива они переставали помещаться по ширине. Слайдер в строке
        # один, его назначение объясняет подсказка.
        hl.addWidget(lbl)
        hl.addWidget(_sep())
        hl.addWidget(alpha_slider, 1)
        hl.addWidget(_sep())
        hl.addWidget(points_btn)

        self._style_ch_row(ch_id)
        return row

    def _style_ch_row(self, ch_id: int):
        """Перекрасить строку канала в его цвет (и приглушить, если канал скрыт).

        Вызывается при создании строки, смене цвета и включении/выключении
        канала — цвет строки всегда совпадает с цветом кривой на графике.
        """
        ch = self._channels.get(ch_id)
        if ch is None or ch.get('row') is None:
            return
        on    = ch['visible']
        color = ch['color'] if on else "#7f8c8d"
        ch['row'].setStyleSheet(f"""
            QFrame#chRow {{
                background: {_rgba(color, 26 if on else 14)};
                border: 1px solid {_rgba(color, 80 if on else 45)};
                border-left: 3px solid {color};
                border-radius: 5px;
            }}
            QFrame#chRow QLabel#chName {{
                color: {color}; font-weight: 600;
                background: transparent;
            }}
            QFrame#chRow QPushButton#chPoints {{
                background: transparent; border: none; border-radius: 4px;
            }}
            QFrame#chRow QPushButton#chPoints:hover {{ background: {_rgba(color, 60)}; }}
            QFrame#chRow QPushButton#chPoints:checked {{ background: {_rgba(color, 90)}; }}
            QFrame#chRow QSlider::handle:horizontal {{ background: {color}; }}
            QFrame#chRow QSlider::handle:horizontal:hover {{
                background: {QColor(color).lighter(125).name()};
            }}
        """)
        pts = ch.get('points_btn')
        if pts is not None:
            pts.setIcon(QIcon(make_icon(
                "points" if ch['points'] else "points_off", color, 17)))
            pts.setToolTip("Скрыть точки на линии" if ch['points']
                           else "Показать точки на линии")

    # ── per-channel controls ──────────────────────────────────────────────────

    def _row_menu(self, ch_id: int, pos):
        """Меню строки канала по правой кнопке: цвет линии и показ на графике."""
        ch = self._channels.get(ch_id)
        if ch is None:
            return
        menu = QMenu(ch['row'])
        menu.addAction("Цвет линии…", lambda cid=ch_id: self._pick_channel_color(cid))
        menu.addAction("Скрыть канал" if ch['visible'] else "Показать канал",
                       lambda cid=ch_id, v=not ch['visible']:
                           self._toggle_ch_visible(cid, v))
        menu.exec(ch['row'].mapToGlobal(pos))

    def _toggle_ch_visible(self, ch_id: int, visible: bool):
        """Показать или скрыть канал на графике, подключив/отключив сигнал."""
        ch = self._channels.get(ch_id)
        if ch is None:
            return
        ch['visible'] = visible
        ch['curve'].setVisible(visible and self._mode == "live")
        if visible and self._mode == "live":
            self._connect_ch(ch_id, ch)
            self._legend.addItem(ch['curve'], ch['name'])
        else:
            self._disconnect_ch(ch)
            self._legend.removeItem(ch['curve'])
        self._style_ch_row(ch_id)

    def _pick_channel_color(self, ch_id: int):
        """Открыть диалог выбора цвета для канала."""
        ch = self._channels.get(ch_id)
        if ch is None:
            return
        color = QColorDialog.getColor(QColor(ch['color']), self)
        if color.isValid():
            self.set_curve_color(ch_id, color.name())

    def _set_ch_width(self, ch_id: int, width: int):
        ch = self._channels.get(ch_id)
        if ch is None:
            return
        ch['width'] = width
        vb = self.plot_widget.getViewBox()
        saved = vb.viewRange()
        ch['curve'].opts['antialias'] = True
        ch['curve'].setPen(pg.mkPen(color=_ch_qcolor(ch), width=width))
        vb.setRange(xRange=saved[0], yRange=saved[1], padding=0)
        self._restyle_archive_line(ch_id)

    def _set_ch_alpha(self, ch_id: int, value: int):
        ch = self._channels.get(ch_id)
        if ch is None:
            return
        ch['alpha'] = int(value * 2.55)
        vb = self.plot_widget.getViewBox()
        saved = vb.viewRange()
        ch['curve'].setPen(pg.mkPen(color=_ch_qcolor(ch), width=ch['width']))
        vb.setRange(xRange=saved[0], yRange=saved[1], padding=0)
        self._restyle_archive_line(ch_id)

    def _set_ch_points(self, ch_id: int, checked: bool):
        ch = self._channels.get(ch_id)
        if ch is None:
            return
        ch['points'] = checked
        vb = self.plot_widget.getViewBox()
        saved = vb.viewRange()
        ch['curve'].setSymbol('o' if checked else None)
        ch['curve'].setSymbolSize(5 if checked else 1)
        ch['curve'].setSymbolBrush(pg.mkBrush(ch['color']) if checked else None)
        vb.setRange(xRange=saved[0], yRange=saved[1], padding=0)
        self._restyle_archive_line(ch_id)
        self._style_ch_row(ch_id)   # значок кнопки показывает текущее состояние

    # ── режимы ────────────────────────────────────────────────────────────────

    def _sync_mode_btns(self):
        """Отметить сегмент текущего режима; значок — в цвет надписи сегмента."""
        for mode, (btn, kind) in self._mode_btns.items():
            on = (mode == self._mode)
            btn.setChecked(on)
            btn.setIcon(QIcon(make_icon(kind, "#ffffff" if on else _MODE_COLORS[mode], 14)))

    def _set_mode(self, mode: str):
        if mode == self._mode:
            self._sync_mode_btns()   # повторный клик по своему же сегменту
            return
        self._mode = mode
        vb = self.plot_widget.getViewBox()
        self._export_menu_action.setVisible(mode == "archive")
        self._sync_mode_btns()
        if mode == "live":
            self._arch_panel.setVisible(False)
            self.plot_widget.setLabel("bottom", "Время")
            self.plot_widget.setLabel("left", "Значение")
            self._clear_archive()
            for ch_id, ch in self._channels.items():
                ch['write'] = 0
                ch['full']  = False
                ch['curve'].setVisible(ch['visible'])
                if ch['visible']:
                    self._connect_ch(ch_id, ch)
                    self._legend.addItem(ch['curve'], ch['name'])
            self._y_range     = (0.0, 1.0)
            self._auto_scroll = True
            # момент включения live — к нему прижат левый край, пока окно не заполнилось
            self._live_start_ts = QDateTime.currentDateTimeUtc().toMSecsSinceEpoch() / 1000.0
            vb.disableAutoRange()
            vb.setAutoVisible(y=False)
            self._render_timer.start(LIVE_RENDER_MS)
        else:
            self._arch_panel.setVisible(True)
            for ch in self._channels.values():
                self._disconnect_ch(ch)
                self._legend.removeItem(ch['curve'])
            self._stop_archive_workers()
            self._render_timer.stop()
            self._clear_archive()
            for ch in self._channels.values():
                ch['curve'].setData([], [])
                ch['curve'].setVisible(False)
            vb.disableAutoRange()
            vb.enableAutoRange()

    def _connect_ch(self, ch_id: int, ch: dict):
        """Подключить сигнал канала (если привязан и ещё не подключён)."""
        if ch['signal'] is None or ch['slot'] is not None:
            return
        def slot(name, times, values, _cid=ch_id, _stream=ch['stream']):
            if name == _stream:
                self._push_points(_cid, times, values)
        ch['slot'] = slot
        # QueuedConnection обязателен: сигналы летят из _db_thread,
        # а кольцевой буфер читается из главного потока в _render_frame.
        # Без явного QueuedConnection лямбда вызывается из _db_thread → race condition.
        ch['signal'].connect(slot, Qt.ConnectionType.QueuedConnection)

    def _disconnect_ch(self, ch: dict):
        """Отключить сигнал канала, сохраняя привязку (signal остаётся)."""
        if ch['slot'] is None or ch['signal'] is None:
            return
        try:
            ch['signal'].disconnect(ch['slot'])
        except RuntimeError:
            pass
        ch['slot'] = None

    # ── live ──────────────────────────────────────────────────────────────────

    def _on_stream_rate(self, name: str, step_ms: float):
        """Воркер сообщил шаг между точками потока → пересчитать ёмкость буферов.

        Темп задаётся ПЛК (max_array_length) и периодом опроса, поэтому ёмкость
        считаем от него, а не фиксируем: буфер должен вмещать всё окно
        LIVE_WINDOW_SECS, иначе старая часть графика молча пропадает.
        Размер берём по самому быстрому потоку — буферы у каналов общей длины.
        """
        if step_ms and step_ms > 0:
            self._stream_step_ms[name] = step_ms
        if not self._stream_step_ms:
            return
        fastest = min(self._stream_step_ms.values())
        need    = int(LIVE_WINDOW_SECS * 1000 / fastest * LIVE_POINTS_MARGIN)
        self._set_live_capacity(max(1000, min(need, LIVE_POINTS_LIMIT)))

    def _set_live_capacity(self, cap: int):
        """Расширить кольцевые буферы каналов до ёмкости cap.

        Только увеличиваем: измеренный темп слегка плавает, и сжатие на каждом
        колебании выбрасывало бы часть уже накопленного графика. Накопленное
        содержимое переносится, чтобы расширение не сбрасывало живую кривую.
        """
        if cap <= self._live_cap:
            return
        self._live_cap = cap
        for ch_id, ch in self._channels.items():
            x, y = self._get_buf_data(ch_id)      # текущее содержимое по порядку
            buf_t = np.empty(cap, dtype=np.float64)
            buf_v = np.empty(cap, dtype=np.float64)
            n = min(len(x), cap)
            if n:
                buf_t[:n] = x[-n:]
                buf_v[:n] = y[-n:]
            ch['buf_t'], ch['buf_v'] = buf_t, buf_v
            ch['write'] = n % cap
            ch['full']  = n == cap

    def _push_points(self, ch_id: int, times: list, values: list):
        """Записать точки в кольцевой буфер канала и обновить кривую."""
        if self._mode != "live":
            return
        ch = self._channels.get(ch_id)
        if ch is None:
            return
        buf_t = ch['buf_t']
        buf_v = ch['buf_v']
        for t, v in zip(times, values):
            w = ch['write']
            buf_t[w] = t
            buf_v[w] = v
            ch['write'] = (w + 1) % len(buf_t)
            if ch['write'] == 0:
                ch['full'] = True
        if values:
            ch['last_val'] = values[-1]

    def _get_buf_data(self, ch_id: int) -> tuple:
        """Вернуть (x, y) из кольцевого буфера канала в хронологическом порядке."""
        ch    = self._channels[ch_id]
        buf_t = ch['buf_t']
        buf_v = ch['buf_v']
        w     = ch['write']
        if ch['full']:
            x = np.concatenate([buf_t[w:], buf_t[:w]])
            y = np.concatenate([buf_v[w:], buf_v[:w]])
        else:
            x = buf_t[:w].copy()
            y = buf_v[:w].copy()
        return x, y

    def _render_frame(self):
        """50 мс: плавный скролл X-оси + обновление Y-диапазона."""
        now_ts = QDateTime.currentDateTimeUtc().toMSecsSinceEpoch() / 1000.0

        # Нормализация X: вычитаем ts_offset чтобы OpenGL работал с малыми числами
        # (float32 имеет точность ±256 сек на значении ~1.7e9 — видимая вибрация)
        if self._auto_scroll:
            # Пока окно не заполнилось, левый край прижат к моменту включения live —
            # перо идёт слева направо, как на самописце. Когда данные дошли до правого
            # края, окно начинает ехать за now. Ось хронологическая в обоих случаях:
            # слева старое, справа новое (раньше левый край всегда был now-300, поэтому
            # свежий график строился от правого края влево).
            self._ts_offset = (self._live_start_ts
                               if now_ts - self._live_start_ts < LIVE_WINDOW_SECS
                               else now_ts - LIVE_WINDOW_SECS)
            self._time_axis.ts_offset = self._ts_offset
            self.plot_widget.setXRange(0, LIVE_WINDOW_SECS, padding=0)

        x_ref    = self._ts_offset
        now_norm = now_ts - x_ref   # «сейчас» в нормализованных координатах

        # обновить все каналы синхронно
        for ch_id, ch in self._channels.items():
            if not ch['visible']:
                continue
            if ch['hold']:
                # hold-канал: шаг-функция от window_start до now_norm
                if ch['last_val'] is None:
                    continue
                win_start_norm = now_norm - LIVE_WINDOW_SECS
                if ch['full'] or ch['write'] > 0:
                    x, y = self._get_buf_data(ch_id)
                    x = x - x_ref
                    mask_before = x <= win_start_norm
                    if np.any(mask_before):
                        # есть точки до окна — продлить от win_start_norm
                        val_at_start = float(y[np.where(mask_before)[0][-1]])
                        mask_in = ~mask_before
                        x_win = np.concatenate([[win_start_norm], x[mask_in], [now_norm]])
                        y_win = np.concatenate([[val_at_start], y[mask_in], [ch['last_val']]])
                    else:
                        # все точки внутри окна — начинать с первой реальной точки
                        x_win = np.concatenate([x, [now_norm]])
                        y_win = np.concatenate([y, [ch['last_val']]])
                else:
                    x_win = np.array([win_start_norm, now_norm])
                    y_win = np.array([ch['last_val'], ch['last_val']])
                ch['curve'].setData(x_win, y_win)
            else:
                # обычный канал: данные из буфера, нормализованные
                if ch['full'] or ch['write'] > 0:
                    x, y = self._get_buf_data(ch_id)
                    ch['curve'].setData((x - x_ref).astype(np.float32),
                                        y.astype(np.float32))

        # Y-диапазон по всем активным и видимым буферам
        all_y = []
        for ch in self._channels.values():
            if not ch['visible']:
                continue
            n = len(ch['buf_v']) if ch['full'] else ch['write']
            if n > 0:
                all_y.append(ch['buf_v'][:n])
        if not all_y:
            return
        y = np.concatenate(all_y)
        ymin, ymax = float(y.min()), float(y.max())
        span = (ymax - ymin) or abs(ymax) or 1.0
        lo, hi = ymin - span * 0.05, ymax + span * 0.05
        prev_lo, prev_hi = self._y_range
        prev_span = (prev_hi - prev_lo) or 1.0
        if abs(lo - prev_lo) / prev_span > 0.05 or abs(hi - prev_hi) / prev_span > 0.05:
            self.plot_widget.getViewBox().setYRange(lo, hi, padding=0)
            self._y_range = (lo, hi)

    def _on_manual_pan(self):
        """Пользователь потащил график мышью — останавливаем авто-скролл."""
        if self._mode == "live":
            self._auto_scroll = False

    def _resume_autoscroll(self):
        if self._mode == "live":
            self._auto_scroll = True
            self._y_range = (0.0, 1.0)

    def _pan_left(self):
        if self._mode != "live":
            return
        vb = self.plot_widget.getViewBox()
        x0, x1 = vb.viewRange()[0]
        step = (x1 - x0) * 0.2
        self._auto_scroll = False
        self.plot_widget.setXRange(x0 - step, x1 - step, padding=0)

    def _pan_right(self):
        if self._mode != "live":
            return
        vb = self.plot_widget.getViewBox()
        x0, x1 = vb.viewRange()[0]
        step = (x1 - x0) * 0.2
        self._auto_scroll = False
        self.plot_widget.setXRange(x0 + step, x1 + step, padding=0)

    # ── архив ─────────────────────────────────────────────────────────────────

    # ch_id → (measurement, field) для запросов архива
    _ARCHIVE_SOURCES = {
        1: ("nowSetpoint",  "value"),
        2: ("tenza",        "value"),
        3: ("displacement", "value"),
    }

    # Прореживание архива на стороне Influx: сколько точек максимум тянуть и
    # рисовать на канал. Экран всё равно не покажет больше — на 1900 точек
    # приходится примерно по точке на пиксель ширины графика, а из базы при
    # этом едут килобайты вместо десятков мегабайт.
    _ARCHIVE_MAX_PTS = 1900

    # Каким каналам прореживание применять и какой функцией сворачивать окно.
    # Уставка (канал 1) сюда не входит: это ступенчатый сигнал с редкими
    # изменениями, усреднение размыло бы ступени, а точек там и так мало.
    _ARCHIVE_AGG = {
        2: "mean",   # нагрузка
        3: "mean",   # перемещение
    }

    def _set_load_progress(self, fraction):
        """fraction=0..1 — заливка кнопки; None — сброс (загрузка завершена)."""
        if fraction is None:
            self.btn_load.setEnabled(True)
            self.btn_load.setText("Загрузить")
            self.btn_load.setStyleSheet("")
            return
        self.btn_load.setEnabled(False)
        pct = int(fraction * 100)
        stop = max(0.0, min(fraction, 1.0))
        self.btn_load.setText(f"Загрузка {pct}%")
        self.btn_load.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2980b9, stop:{stop:.3f} #2980b9,
                    stop:{min(stop + 0.001, 1.0):.3f} #555555, stop:1 #555555);
                color: white;
                border: none;
                border-radius: 3px;
            }}
        """)

    def _load_archive(self):
        qdt_from = QDateTime(self.date_from.date(), self.time_from.time()).toUTC()
        qdt_to   = QDateTime(self.date_to.date(),   self.time_to.time()).toUTC()

        # нормализация X для архива: отсчёт от начала запрошенного диапазона
        self._ts_offset = qdt_from.toMSecsSinceEpoch() / 1000.0
        self._time_axis.ts_offset = self._ts_offset

        self._clear_archive()          # сбрасывает и _arch_full — выставляем после
        # весь заказанный диапазон: за его пределы перезапрос при зуме не выйдет
        self._arch_full = (qdt_from.toMSecsSinceEpoch() / 1000.0,
                           qdt_to.toMSecsSinceEpoch() / 1000.0)
        self.plot_widget.getViewBox().enableAutoRange()   # первый показ — по данным
        self._query_archive(*self._arch_full)

    def _query_archive(self, t_from: float, t_to: float):
        """Запросить архив за [t_from, t_to] (секунды Unix) и обновить кривые.

        Каждый канал прореживается в самой базе до _ARCHIVE_MAX_PTS точек, а не
        тянется целиком: за сутки в базе миллионы отсчётов, на экране их всё
        равно не различить, зато передача и отрисовка занимали десятки секунд.
        """
        if t_to <= t_from:
            return
        self._stop_archive_workers()
        self._archive_parts.clear()
        self._workers_done = 0

        # видимые каналы для загрузки
        visible_sources = {
            ch_id: src
            for ch_id, src in self._ARCHIVE_SOURCES.items()
            if self._channels.get(ch_id, {}).get('visible', True)
        }
        if not visible_sources:
            return
        self._arch_window = (t_from, t_to)
        self._total_archive_workers = len(visible_sources) * N_ARCHIVE_WORKERS
        self._set_load_progress(0)

        # Окно свёртки — на весь диапазон, а не на кусок воркера: части идут
        # подряд с одинаковым шагом, поэтому в сумме выходит ровно тот предел,
        # что задан на канал.
        span   = t_to - t_from
        every  = max(1_000_000, int(span * 1e9 / self._ARCHIVE_MAX_PTS))
        step   = span / N_ARCHIVE_WORKERS

        for ch_id, (measurement, field) in visible_sources.items():
            self._archive_parts[ch_id] = {}
            agg = self._ARCHIVE_AGG.get(ch_id)
            for i in range(N_ARCHIVE_WORKERS):
                p_from = t_from + i * step
                p_to   = t_from + (i + 1) * step if i < N_ARCHIVE_WORKERS - 1 else t_to
                # keep — только нужные колонки (меньше байт по HTTP и парсинга);
                # group()/sort() не нужны: один воркер тянет непрерывный кусок
                # одной серии, Influx отдаёт его по времени по возрастанию.
                query = f'''
from(bucket: "{INFLUX_BUCKET}")
  |> range(start: {_flux_time(p_from)}, stop: {_flux_time(p_to)})
  |> filter(fn: (r) => r._measurement == "{measurement}" and r._field == "{field}")
  |> keep(columns: ["_time", "_value"])
'''
                if agg:
                    query += (f'  |> aggregateWindow(every: {every}ns, '
                              f'fn: {agg}, createEmpty: false)\n')
                worker = _ArchiveWorker(i, query, parent=self)
                worker.part_ready.connect(
                    lambda idx, t, v, c=ch_id: self._on_archive_part(c, idx, t, v)
                )
                worker.finished.connect(self._on_archive_worker_done)
                self._archive_workers.append(worker)

        for w in self._archive_workers:
            w.start()

    def _stop_archive_workers(self):
        """Отцепить сигналы прежних воркеров — их ответы уже не нужны."""
        for w in self._archive_workers:
            try:
                w.part_ready.disconnect()
            except (RuntimeError, TypeError):
                pass
            try:
                w.finished.disconnect(self._on_archive_worker_done)
            except (RuntimeError, TypeError):
                pass
        self._archive_workers.clear()

    # ── перезапрос при смене масштаба ─────────────────────────────────────────

    def _on_arch_x_changed(self, _vb=None, _rng=None):
        """Масштаб или сдвиг по X изменились — перечитать архив под новое окно.

        Сразу не дёргаем: при перетаскивании сигнал приходит десятки раз в
        секунду, запрос уходит только когда пользователь остановился.
        """
        if self._mode != "archive" or self._arch_full is None:
            return
        self._arch_timer.start()

    def _visible_archive_range(self):
        """Видимое окно в секундах Unix, обрезанное по загруженному диапазону."""
        (x0, x1), _ = self.plot_widget.getViewBox().viewRange()
        t0 = max(self._arch_full[0], x0 + self._ts_offset)
        t1 = min(self._arch_full[1], x1 + self._ts_offset)
        return t0, t1

    def _reload_visible_archive(self):
        """Перечитать архив под текущее видимое окно, если оно заметно съехало."""
        if self._mode != "archive" or self._arch_full is None:
            return
        t0, t1 = self._visible_archive_range()
        if t1 - t0 <= 0:
            return
        if self._arch_window is not None:
            w0, w1 = self._arch_window
            span = max(w1 - w0, 1e-9)
            # окно почти то же самое — перезапрашивать нечего
            if abs(t0 - w0) / span < 0.02 and abs(t1 - w1) / span < 0.02:
                return
        # вид уже выставлен пользователем: новые данные не должны его двигать
        self.plot_widget.getViewBox().disableAutoRange()
        self._query_archive(t0, t1)

    def _on_archive_part(self, ch_id: int, idx: int, times, values):
        """Часть канала пришла (times/values — numpy-массивы от воркера)."""
        if ch_id not in self._archive_parts:
            return
        self._archive_parts[ch_id][idx] = (times, values)
        # ждём ВСЕ части канала, затем строим кривую один раз —
        # без промежуточных setData (каждый из них пересчитывал даунсэмпл по всем точкам)
        if len(self._archive_parts[ch_id]) < N_ARCHIVE_WORKERS:
            return
        parts_t, parts_v = [], []
        for i in range(N_ARCHIVE_WORKERS):
            t, v = self._archive_parts[ch_id][i]
            if len(t):
                parts_t.append(t)
                parts_v.append(v)
        if not parts_t:
            return
        # части идут в порядке idx 0…N-1; интервалы воркеров не пересекаются,
        # а внутри куска сортировка уже сделана в воркере → склейки достаточно,
        # полный argsort по всему массиву на GUI-потоке не нужен (он и морозил интерфейс)
        import time as _t                       # DEBUG-тайминг: убрать после диагностики
        _g0 = _t.perf_counter()
        x = np.concatenate(parts_t)
        y = np.concatenate(parts_v)
        _g1 = _t.perf_counter()
        self._update_archive_curve(ch_id, x, y)
        _g2 = _t.perf_counter()
        print(f"[GUI ch{ch_id}] точек={len(x):>9}  склейка={_g1 - _g0:6.3f}с  "
              f"setData/отрисовка={_g2 - _g1:6.3f}с")

    def _on_archive_worker_done(self):
        self._workers_done += 1
        self._set_load_progress(self._workers_done / max(self._total_archive_workers, 1))
        if self._workers_done >= self._total_archive_workers:
            self._archive_workers.clear()
            self._set_load_progress(None)

    def _update_archive_curve(self, ch_id: int, x, y):
        ch       = self._channels.get(ch_id, {})
        if ch.get('hold', False) and len(x) > 1:
            x, y = _step_xy(x, y)
        x_norm = x - self._ts_offset   # нормализация для OpenGL
        color    = ch.get('color',  "#e67e22")
        width    = ch.get('width',  _CH_WIDTH)
        show_pts = ch.get('points', False)
        name = ch.get('name', f"канал {ch_id}")
        pen  = pg.mkPen(color=color, width=width)
        if ch_id in self._archive_lines:
            self._archive_lines[ch_id].setData(x_norm, y)
        else:
            line = self.plot_widget.plot(
                x_norm, y, pen=pen,
                symbol='o' if show_pts else None,
                symbolSize=5 if show_pts else 1,
                symbolBrush=pg.mkBrush(color) if show_pts else None,
            )
            line.setDownsampling(auto=True, method='subsample')
            line.setClipToView(True)
            self._archive_lines[ch_id] = line
            self._legend.addItem(line, name)
        # оформление могли поменять, пока архив грузился
        self._restyle_archive_line(ch_id)

    def _restyle_archive_line(self, ch_id: int):
        """Применить настройки канала к его архивной кривой.

        Цвет, толщина, прозрачность и точки правят live-кривую, а архивная —
        отдельный объект графика. Без этого переключатели панели действовали
        только на живой режим, а загруженный архив оставался таким, каким был
        в момент загрузки.
        """
        line = self._archive_lines.get(ch_id)
        ch   = self._channels.get(ch_id)
        if line is None or ch is None:
            return
        line.setPen(pg.mkPen(color=_ch_qcolor(ch), width=ch['width']))
        on = ch['points']
        line.setSymbol('o' if on else None)
        line.setSymbolSize(5 if on else 1)
        line.setSymbolBrush(pg.mkBrush(_ch_qcolor(ch)) if on else None)

    # ── перекрестие и подсказка ───────────────────────────────────────────────

    def _on_mouse_moved(self, pos):
        rect = self.plot_widget.sceneBoundingRect()
        if not rect.contains(pos):
            self._vline.setVisible(False)
            self._hline.setVisible(False)
            return
        mp = self.plot_widget.getPlotItem().vb.mapSceneToView(pos)
        self._vline.setPos(mp.x())
        self._hline.setPos(mp.y())
        self._vline.setVisible(True)
        self._hline.setVisible(True)

    def _on_mouse_clicked(self, event):
        if event.button() == Qt.MouseButton.RightButton:
            self._click_marker.setVisible(False)
            self._click_label.setVisible(False)
            return
        if event.double():
            self._resume_autoscroll()
            return
        if self._mode == "live":
            self._auto_scroll = False
        pos = event.scenePos()
        if not self.plot_widget.sceneBoundingRect().contains(pos):
            return
        mp      = self.plot_widget.getPlotItem().vb.mapSceneToView(pos)
        click_x = mp.x()
        click_y = mp.y()

        best_ch_name, best_ch_xs, best_ch_ys = None, None, None
        best_y_dist = float('inf')
        for name, xs, ys in self._iter_curve_data():
            if len(xs) == 0:
                continue
            fi = int(np.searchsorted(xs, click_x, side='right')) - 1
            fi = np.clip(fi, 0, len(xs) - 1)
            y_at_click = float(ys[fi])
            if np.isnan(y_at_click):
                continue
            d = abs(y_at_click - click_y)
            if d < best_y_dist:
                best_y_dist = d
                best_ch_name, best_ch_xs, best_ch_ys = name, xs, ys

        if best_ch_name is None:
            return
        _, (y0, y1) = self.plot_widget.getViewBox().viewRange()
        y_span = (y1 - y0) or 1.0
        if best_y_dist / y_span > 0.05:
            return

        ni = int(np.searchsorted(best_ch_xs, click_x))
        ni = np.clip(ni, 0, len(best_ch_xs) - 1)
        if ni > 0 and abs(best_ch_xs[ni - 1] - click_x) < abs(best_ch_xs[ni] - click_x):
            ni -= 1
        best_x, best_y, best_name = float(best_ch_xs[ni]), float(best_ch_ys[ni]), best_ch_name
        time_str = _dt.datetime.fromtimestamp(best_x + self._ts_offset).strftime("%H:%M:%S.%f")[:-3]
        self._click_marker.setData([best_x], [best_y])
        self._click_label.setText(f"{best_name}\n{time_str}\n{best_y:.4f}")
        self._click_label.setPos(best_x, best_y)
        self._click_marker.setVisible(True)
        self._click_label.setVisible(True)
        event.accept()

    def _iter_curve_data(self):
        """Итератор (name, xs, ys) по всем видимым кривым."""
        if self._mode == "live":
            for ch in self._channels.values():
                if not ch['visible']:
                    continue
                xs, ys = ch['curve'].getData()
                if xs is not None and len(xs) > 0:
                    yield ch['name'], xs, ys
        else:
            for ch_id, line in self._archive_lines.items():
                xs, ys = line.getData()
                if xs is not None and len(xs) > 0:
                    yield self._channels[ch_id]['name'], xs, ys

    # ── контекстное меню графика ──────────────────────────────────────────────

    def set_theme(self, dark: bool):
        axis_color = "#aaaaaa" if dark else "#333333"
        plot_bg    = "#2c2c2c" if dark else "#f5f5f5"
        self.plot_widget.setBackground(plot_bg)
        for axis in ("left", "bottom"):
            self.plot_widget.getAxis(axis).setPen(pg.mkPen(axis_color))
            self.plot_widget.getAxis(axis).setTextPen(pg.mkPen(axis_color))

        panel_ss = _panel_style(dark)
        self._nav_frame.setStyleSheet(_nav_style(dark))
        self._ch_frame.setStyleSheet(panel_ss)
        self._arch_panel.setStyleSheet(panel_ss)

    def _build_graph_context_menu(self):
        menu = self.plot_widget.getViewBox().menu
        menu.addSeparator()

        # настройки линий перенесены в панель каналов

        grid_menu = menu.addMenu("Сетка")
        act_grid_x = QAction("По X", grid_menu)
        act_grid_x.setCheckable(True)
        act_grid_x.setChecked(True)
        act_grid_y = QAction("По Y", grid_menu)
        act_grid_y.setCheckable(True)
        act_grid_y.setChecked(True)

        grid_alpha_val = [0.3]

        def _update_grid():
            self.plot_widget.showGrid(
                x=act_grid_x.isChecked(), y=act_grid_y.isChecked(),
                alpha=grid_alpha_val[0])

        act_grid_x.toggled.connect(lambda _: _update_grid())
        act_grid_y.toggled.connect(lambda _: _update_grid())
        grid_menu.addAction(act_grid_x)
        grid_menu.addAction(act_grid_y)

        grid_menu.addSeparator()
        alpha_gw = QWidget()
        alpha_gl = QHBoxLayout(alpha_gw)
        alpha_gl.setContentsMargins(16, 4, 8, 4)
        alpha_gl.setSpacing(6)
        alpha_gl.addWidget(QLabel("Прозрачность:"))
        grid_slider = QSlider(Qt.Orientation.Horizontal)
        grid_slider.setRange(0, 100)
        grid_slider.setValue(30)
        grid_slider.setFixedWidth(90)
        def _on_grid_alpha(v):
            grid_alpha_val[0] = v / 100.0
            _update_grid()
        grid_slider.valueChanged.connect(_on_grid_alpha)
        alpha_gl.addWidget(grid_slider)
        alpha_ga = QWidgetAction(grid_menu)
        alpha_ga.setDefaultWidget(alpha_gw)
        grid_menu.addAction(alpha_ga)

    # ── вспомогательное ───────────────────────────────────────────────────────

    def _clear_archive(self):
        for line in self._archive_lines.values():
            self._legend.removeItem(line)
            self.plot_widget.removeItem(line)
        self._archive_lines.clear()
        self._archive_parts.clear()
        self._arch_full = self._arch_window = None

    # ── публичный API ─────────────────────────────────────────────────────────

    def bind_channel(self, ch_id: int, signal) -> None:
        """Привязать pyqtSignal к каналу ch_id (1, 2 или 3).

        signal должен иметь сигнатуру (times: list, values: list).
        Предыдущая привязка снимается автоматически.
        """
        ch = self._channels.get(ch_id)
        if ch is None:
            return
        self.unbind_channel(ch_id)
        ch['signal'] = signal
        if self._mode == "live" and ch['visible']:
            self._connect_ch(ch_id, ch)

    def unbind_channel(self, ch_id: int) -> None:
        """Снять привязку сигнала от канала ch_id."""
        ch = self._channels.get(ch_id)
        if ch is None:
            return
        self._disconnect_ch(ch)
        ch['signal'] = None

    def set_curve_color(self, ch_id: int, color: str) -> None:
        """Изменить цвет линии канала ch_id.

        Пример:
            trends.set_curve_color(1, "#ff0000")
        """
        ch = self._channels.get(ch_id)
        if ch is None:
            return
        ch['color'] = color
        ch['curve'].setPen(pg.mkPen(color=_ch_qcolor(ch), width=ch['width']))
        self._restyle_archive_line(ch_id)
        self._style_ch_row(ch_id)   # строка канала красится вместе с кривой

    def set_labels(self, x_label="X", y_label="Y"):
        self.plot_widget.setLabel("bottom", x_label)
        self.plot_widget.setLabel("left", y_label)

    def save_plot(self, filename):
        exporter = pg.exporters.ImageExporter(self.plot_widget.plotItem)
        exporter.export(filename)

    # ── экспорт ───────────────────────────────────────────────────────────────

    def _ask_export_channels(self) -> list[int]:
        """Диалог выбора каналов для выгрузки. Возвращает список ch_id или []."""
        available = {
            ch_id: self._channels[ch_id]['name']
            for ch_id in self._archive_lines
            if ch_id in self._channels
        }
        if not available:
            return []

        dlg = QDialog(self)
        dlg.setWindowTitle("Выгрузка данных")
        dlg.setMinimumWidth(320)
        layout = QVBoxLayout(dlg)
        layout.addWidget(QLabel("Выберите каналы для выгрузки:"))

        checks: dict[int, QCheckBox] = {}
        for ch_id, name in available.items():
            cb = QCheckBox(name)
            cb.setChecked(True)
            checks[ch_id] = cb
            layout.addWidget(cb)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        layout.addWidget(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return []
        return [ch_id for ch_id, cb in checks.items() if cb.isChecked()]

    def _export_csv(self):
        selected = self._ask_export_channels()
        if not selected:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить CSV", "", "CSV файлы (*.csv)")
        if not path:
            return

        def _fmt(ts):
            dt = _dt.datetime.fromtimestamp(ts).astimezone()
            return dt.strftime("%d.%m.%Y %H:%M:%S.%f")[:-3]

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            for ch_id in selected:
                xs, ys = self._archive_lines[ch_id].getData()
                if xs is None or len(xs) == 0:
                    continue
                name = self._channels[ch_id]['name']
                f.write(f"# {name}\n")
                f.write("Время,Значение\n")
                for t, v in zip(xs, ys):
                    f.write(f"{_fmt(t)},{round(float(v), 6)}\n")
                f.write("\n")

    def _export_excel(self):
        selected = self._ask_export_channels()
        if not selected:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить Excel", "", "Excel файлы (*.xlsx)")
        if not path:
            return

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from influxdb_client import InfluxDBClient
        from ._archive_worker import INFLUX_URL, INFLUX_TOKEN, INFLUX_ORG, INFLUX_BUCKET

        qdt_from = QDateTime(self.date_from.date(), self.time_from.time()).toUTC()
        qdt_to   = QDateTime(self.date_to.date(),   self.time_to.time()).toUTC()
        t_from   = qdt_from.toString("yyyy-MM-ddTHH:mm:ssZ")
        t_to     = qdt_to  .toString("yyyy-MM-ddTHH:mm:ssZ")

        self.btn_load.setEnabled(False)
        self.btn_load.setText("Экспорт…")

        # прямой запрос к InfluxDB для каждого выбранного канала
        ch_data: list[tuple[str, np.ndarray, np.ndarray]] = []
        try:
            client = InfluxDBClient(url=INFLUX_URL, token=INFLUX_TOKEN, org=INFLUX_ORG,
                                    enable_gzip=True)
            for ch_id in selected:
                measurement, field = self._ARCHIVE_SOURCES[ch_id]
                query = f'''
from(bucket: "{INFLUX_BUCKET}")
  |> range(start: {t_from}, stop: {t_to})
  |> filter(fn: (r) => r._measurement == "{measurement}" and r._field == "{field}")
  |> group()
  |> sort(columns: ["_time"])
'''
                times, values = [], []
                for rec in client.query_api().query_stream(query):
                    times.append(rec.get_time().timestamp())
                    values.append(rec.get_value())
                if times:
                    ch_data.append((self._channels[ch_id]['name'],
                                    np.asarray(times,  dtype=np.float64),
                                    np.asarray(values, dtype=np.float64)))
        except Exception as e:
            print(f"[Export] ошибка запроса: {e}")
        finally:
            try:
                client.close()
            except Exception:
                pass
            self.btn_load.setEnabled(True)
            self.btn_load.setText("Загрузить")

        if not ch_data:
            return

        # мастер-сетка: тенза (ch_id=2), иначе канал с наибольшим числом точек
        master_name = self._channels[2]['name'] if 2 in selected else None
        master_ts   = None
        for name, xs, ys in ch_data:
            if name == master_name:
                master_ts = xs
                break
        if master_ts is None:
            master_ts = max(ch_data, key=lambda d: len(d[1]))[1]

        def _ffill(ts_chan, vs_chan, ts_all):
            """Forward-fill — для редких каналов (уставка)."""
            idx    = np.searchsorted(ts_chan, ts_all, side='right') - 1
            result = np.full(len(ts_all), np.nan)
            mask   = idx >= 0
            result[mask] = vs_chan[idx[mask]]
            if len(vs_chan) > 0 and not mask.all() and mask.any():
                first = int(np.searchsorted(ts_all, ts_chan[0], side='left'))
                result[:first] = vs_chan[0]
            return result

        def _nearest(ts_chan, vs_chan, ts_all, threshold=0.008):
            """Nearest-fill с порогом — для каналов с тем же темпом (смещение)."""
            if len(ts_chan) == 0:
                return np.full(len(ts_all), np.nan)
            idx    = np.searchsorted(ts_chan, ts_all)
            prev_i = np.clip(idx - 1, 0, len(ts_chan) - 1)
            next_i = np.clip(idx,     0, len(ts_chan) - 1)
            prev_d = np.abs(ts_chan[prev_i] - ts_all)
            next_d = np.abs(ts_chan[next_i] - ts_all)
            best_i = np.where(prev_d <= next_d, prev_i, next_i)
            best_d = np.minimum(prev_d, next_d)
            result = np.full(len(ts_all), np.nan)
            mask   = best_d <= threshold
            result[mask] = vs_chan[best_i[mask]]
            return result

        # смещение и тенза — один темп (nearest); уставка — редкая (ffill)
        _dense_names = {self._channels[2]['name'], self._channels[3]['name']}
        aligned = [
            (name, _nearest(xs, ys, master_ts) if name in _dense_names else _ffill(xs, ys, master_ts))
            for name, xs, ys in ch_data
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Архив"

        header_fill = PatternFill("solid", fgColor="1A8FE3")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center")

        headers    = ["Дата", "Время"] + [f"Значение {name}" for name, _ in aligned]
        col_widths = [12, 15] + [max(20, len(h) + 2) for h in headers[2:]]
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            ws.column_dimensions[cell.column_letter].width = w

        val_cols = [vals for _, vals in aligned]
        for row, ts in enumerate(master_ts, 2):
            dt = _dt.datetime.fromtimestamp(float(ts)).astimezone()
            ws.cell(row=row, column=1, value=dt.strftime("%d.%m.%Y"))
            ws.cell(row=row, column=2, value=dt.strftime("%H:%M:%S.%f")[:-3])
            for col, vals in enumerate(val_cols, 3):
                v = vals[row - 2]
                ws.cell(row=row, column=col,
                        value=round(float(v), 3) if not np.isnan(v) else None)

        wb.save(path)


# ── вспомогательные функции модуля ───────────────────────────────────────────

def _flux_time(ts: float) -> str:
    """Секунды Unix → метка времени RFC3339 в UTC для запроса Flux."""
    return (_dt.datetime.fromtimestamp(ts, _dt.timezone.utc)
            .strftime("%Y-%m-%dT%H:%M:%S.%fZ"))


def _step_xy(x: np.ndarray, y: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    """Преобразовать массивы точек в ступенчатую функцию (горизонталь + вертикальный прыжок).

    Для N точек возвращает 2N-1 точек:
        [x0, x1, x1, x2, x2, ..., xn-1]
        [y0, y0, y1, y1, ..., yn-1]
    """
    if len(x) <= 1:
        return x, y
    sx = np.repeat(x, 2)[1:]    # [x0, x1, x1, x2, x2, ..., xn-1]
    sy = np.repeat(y, 2)[:-1]   # [y0, y0, y1, y1, ..., yn-1]
    return sx, sy


def _sep() -> QFrame:
    """Вертикальный разделитель для строки канала."""
    # заливкой, а не QFrame.VLine: рамки в строке уже перекрыты её стилем
    sep = QFrame()
    sep.setFixedWidth(1)
    sep.setStyleSheet(f"background: {_rgba('#ffffff', 34)}; border: none;")
    return sep


def _cap(text: str) -> QLabel:
    """Подпись поля в строке канала: мелкая и приглушённая, чтобы не спорить с именем."""
    lbl = QLabel(text)
    lbl.setStyleSheet("color: #8a929c; background: transparent; font-size: 11px;")
    return lbl


def _rgba(hex_color: str, alpha: int) -> str:
    """Цвет, разведённый до полупрозрачного (alpha 0…255).

    Полупрозрачность вместо готового оттенка: подложка подмешивается к тому фону,
    что под ней, и одна и та же строка читается на светлой и тёмной теме.
    """
    c = QColor(hex_color)
    return f"rgba({c.red()}, {c.green()}, {c.blue()}, {alpha})"


def _ch_qcolor(ch: dict) -> QColor:
    """QColor канала с учётом прозрачности."""
    c = QColor(ch['color'])
    c.setAlpha(ch['alpha'])
    return c
